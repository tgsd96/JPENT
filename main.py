#Imports
from flask import Flask,request,send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_restful import Resource, Api, reqparse
from flask_cors import CORS
import os
import uuid
import datetime
import openpyxl as op
import random
from configs import configs
import pandas as pd
import math

#App configs
app = Flask(__name__)
CORS(app)
api = Api(app)
UPLOAD_FOLDER ='C:\\Users\\abcd\\Desktop\\JPENT'
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://tushargarg:colgate@localhost/jpent'
DestFileName = 'result.xlsx'
db = SQLAlchemy(app)


# Database classes
class PurchaseTable(db.Model):
    __tablename__ = 'purchases'
    id = db.Column('id',db.Integer,primary_key=True)
    custid = db.Column('custid',db.Integer)
    billno = db.Column('billno', db.String(200))
    date = db.Column('date', db.Date)
    billvalue = db.Column('billvalue', db.Float)
    interface_code = db.Column('interface_code',db.String(500))

class uploaded(db.Model):
    def __init__(self,location,original,filename,company):
        # self.id = id
        self.location = location
        self.original = original
        self.filename = filename
        self.company = company


    id = db.Column('id', db.Integer, primary_key=True)
    location = db.Column('location', db.String(64))
    original = db.Column('original',db.String(256))
    filename = db.Column('filename', db.String(256))
    company = db.Column('company', db.String(10))
    date = db.Column('date',db.Date,default=datetime.datetime.now().date())

# Main master list table
class MasterList(db.Model):
    __tablename__='master_list'
    cust_id = db.Column('cust_id',db.Integer,primary_key=True)
    uid = db.Column('uid',db.String(500),primary_key=True)
    name = db.Column('name',db.String(500),primary_key=True)
    area = db.Column('area',db.String(500),primary_key=True)
    interface_code = db.Column('interface_code',db.String(500),primary_key=True)

class ErrorPurchases(db.Model):
    id = db.Column('id',db.Integer,primary_key=True)
    custid = db.Column('custid',db.Integer)
    billno = db.Column('billno', db.String(200))
    date = db.Column('date', db.Date)
    billvalue = db.Column('billvalue', db.Float)
    interface_code = db.Column('interface_code',db.String(500))
    name = db.Column('name', db.String(500))


#API classes
class uploadFile(Resource):
    def post(self,company):
        if 'upload' not in request.files:
            response.status = 404
            response.message = "No file given"
            return response
        file = request.files['upload']
        # print(request.form["skipRows"])
        # date = req["date"]
    
        ext = os.path.splitext(file.filename)[1]
        f_name = str(uuid.uuid4())+ext
        newFile = uploaded(UPLOAD_FOLDER,file.filename,f_name,company)
        db.session.add(newFile)
        db.session.commit()
        file.save(os.path.join(UPLOAD_FOLDER,f_name))
        keys = retrieveColumns(os.path.join(UPLOAD_FOLDER,f_name))
        # s_count, e_count = store(os.path.join(UPLOAD_FOLDER,f_name),company)
        # return {
        #     'success' : s_count,
        #     'error' : e_count
        # }
        return {
            'file' : f_name,
            'keys' : keys,
        }
        # return {}
    def put(self,company):
        conf = request.json
        print(conf)
        filename = conf["filename"]
        val = storeWithPandas(filename, conf)
        return { 'status' : val }
    

## download purchases api.
class DownloadPurchase(Resource):
    def get(self,purchase):
        parser = reqparse.RequestParser()
        responseList = []
        parser.add_argument('startDate',type=str)
        parser.add_argument('endDate',type=str)
        dates = parser.parse_args()
        if(purchase=='list'):
            sql = '''
                SELECT COALESCE(S.CUST_ID, BS.CUST_ID) AS CUST_ID, S.cad,S.col,S.god,S.mar,S.marg,S.total, BS.DEBIT AS OLD FROM (
                SELECT A.*,SUM(B.billvalue) total from (
                SELECT * 
                FROM crosstab(
                'select  custid,interface_code,string_agg((billvalue::numeric)::text,''|'')
                FROM PURCHASES 
                WHERE date>='%s'
                and date<='%s'
                group by custid,interface_code order by 1,2','
                SELECT DISTINCT INTERFACE_CODE from master_list order by 1')
                AS ct(cust_id int, CAD text, COL text, GOD text, MAR text, MARG text)) A,
                PURCHASES B where A.cust_id = B.custid
                and B.date>=%s and B.date<=%s
                group by A.cust_id, A.cad, A.col, A.god, A.mar, A.marg) S
                FULL OUTER JOIN
                BALANCE_SHEET BS
                ON S.CUST_ID = BS.CUST_ID
                '''
            result = db.engine.execute(sql,(dates.startDate,dates.endDate,dates.startDate,dates.endDate))
            for row in result:
                # print(row.cust_id);
                getNameQuery = MasterList.query.filter_by(cust_id = row.cust_id).first()
                if getNameQuery == None:
                    print(row.cust_id)
                    continue
                name = getNameQuery.name
                area = getNameQuery.area
                # print(row['cust_id'])
                listList = [
                    name,
                    area,
                    row.col,
                    row.marg,
                    row.god,
                    row.mar,
                    row.cad,
                    row.total,
                    row.old
                ]
                responseList.append(listList)
            wb = op.Workbook()
            ws1 = wb.active
            ws1.title = "List"
            ws1.append(['Name','Beat', 'COLGATE', 'MARG', 'GODREJ','MARICO','CADBURY','Total','OLD'])
            for purchase in responseList:
                ws1.append(purchase)
            wb.save(filename=DestFileName)
        else:
            purchasesList = db.session.query(PurchaseTable).\
                        filter(PurchaseTable.date<=dates.endDate).\
                        filter(PurchaseTable.date>=dates.startDate).all()
            for each in purchasesList:
                getNameQuery = MasterList.query.filter_by(cust_id = each.custid).first()
                name = getNameQuery.name
                purchaseDict = {
                    'id' : str(each.id),
                    'custid' : str(each.custid),
                    'bill_no' : str(each.billno),
                    'interface_code':str(each.interface_code),
                    'date' : str(each.date),
                    'party' : name,
                    'billvalue' : str(each.billvalue),
                }
                responseList.append(purchaseDict)
            wb = op.Workbook()
            ws1 = wb.active
            ws1.title = "All purchases"
            ws1.append(responseList[0].keys())
            for purchase in responseList:
                ws1.append(purchase.values())
            wb.save(filename=DestFileName)
        return send_from_directory(UPLOAD_FOLDER,DestFileName)

class DownloadFile(Resource):
    
    def get(self):
        purchaseList = viewPurchases.get()
        wb = op.Workbook()
        ws1 = wb.active
        ws1.title = "All purchases"

        for purchase in purchaseList:
            ws1.append(purchase)

        wb.save(filename=DestFileName)
        return send_from_directory(UPLOAD_FOLDER,DestFileName)

# resource for Errors
class ErrorHandler(Resource):
    
    def get(self):
        ErrorList = []
        sqlGetErrors = '''
            SELECT distinct name,interface_code
            from error_purchases
        '''
        ErrorNames = db.engine.execute(sqlGetErrors)
        for row in ErrorNames:
            sqlGetRecommendation = '''
                SELECT distinct cust_id, name from master_list where levenshtein(name,%s)<3
            '''
            recommendations = db.engine.execute(sqlGetRecommendation,row.name)
            recommendationArray = []
            for rrow in recommendations:
                recomObject={
                    'cust_id':rrow.cust_id,
                    'name': rrow.name
                }
                recommendationArray.append(recomObject)
            ErrorObject = {
                'name' : row.name,
                'interface_code' : row.interface_code,
                'recommendations' : recommendationArray
            }
            ErrorList.append(ErrorObject)
        return ErrorList

    def post(self):
        data = request.get_json()['data']
        newSql = '''
            Insert into master_list(cust_id, name, interface_code)
            select distinct nextval('cust_id'),A.name,A.interface_code from error_purchases A
            where A.name=%s
        '''
        mergeSQL = '''
            Insert into master_list(cust_id, name, interface_code)
            select distinct %s,A.name,A.interface_code from error_purchases A
            where A.name=%s
        '''
        insertSQL = '''
            Insert into purchases(custid,billno,date,billvalue,interface_code)
            select A.cust_id,B.billno,B.date,B.billvalue,B.interface_code from master_list A, error_purchases B
            where A.name = B.name;
        '''
        deleteSQL='''
            Delete from error_purchases where name in (select name from master_list)
        '''
        for obj in data:
            if(obj['custid']==-1):
                db.engine.execute(newSql,obj['name'])
            else:
                db.engine.execute(mergeSQL,(obj['custid'],obj['name']))
        db.engine.execute(insertSQL)
        db.engine.execute(deleteSQL)
        db.session.commit()
        return "OK"

class AccountHandler(Resource):
    
    def get(self):
        parser = reqparse.RequestParser()
        responseList = []
        parser.add_argument('name',type=str)
        name = parser.parse_args()
        selectSql = '''
            select  cust_id, name, interface_code from master_list where UPPER(name) like %s order by cust_id
        '''
        key = name.name.upper()+"%"
        result = db.engine.execute(selectSql,key)

        for row in result:
            response = {
                'key' : row['name']+str(row['cust_id'])+row['interface_code'],
                'cust_id' : row['cust_id'],
                'name' : row['name'],
                'interface_code': row['interface_code']
            }
            responseList.append(response)

        return responseList
        
class AccountDetails(Resource):
    
    def get(self):
        parser = reqparse.RequestParser()
        parser.add_argument('name',type=str)
        parser.add_argument('cust_id',type=int)
        parser.add_argument('ic',type=str)
        details = parser.parse_args()
        selectSql = '''
        SELECT name,interface_code 
        from master_list where cust_id=%s
        '''
        recommended = db.engine.execute(selectSql,details['cust_id'])
        recommendList = []

        for row in recommended:
            rr = {
                'name' : row['name'],
                'interface_code' : row['interface_code']
            }
            recommendList.append(rr)
        recommendSQL = '''
            SELECT cust_id,name from master_list
            where levenshtein(UPPER(name),UPPER(%s))<3 and cust_id <> %s
        '''
        recommended = db.engine.execute(recommendSQL,(details.name,details.cust_id))
        recommendersList = []
        for row in recommended:
            rr = {
                'name' : row['name'],
                'cust_id': row['cust_id']
            }
            recommendersList.append(rr)
    
        result={
            'cust_id' : details.cust_id,
            'name' : details.name,
            'interface_code' : details.ic,
            'same' : recommendList,
            'recom': recommendersList
        }

        return result

    def put(self):
        old_id = request.get_json()['cust_id']
        new_id = request.get_json()['new_cust_id']
        updateSQL = '''
        Update master_list set cust_id = %s where cust_id = %s
        '''
        updatePurchases='''
        update purchases set custid = %s where custid=%s
        '''
        db.engine.execute(updateSQL,(new_id,old_id))
        db.engine.execute(updatePurchases,(new_id,old_id))
        db.session.commit()
        return "ok"

#api Endpoints
api.add_resource(uploadFile, '/upload/<company>')
api.add_resource(DownloadPurchase,'/download/<purchase>')
api.add_resource(ErrorHandler,'/errors')
api.add_resource(AccountHandler,'/accounts')
api.add_resource(AccountDetails,'/accounts/details')

# Storing function


def retrieveColumns(filename):
    # print("Skip rows are: %d "%int(skipRows))
    sheet = pd.read_excel(filename)
    return list(sheet.keys())

def storeWithPandas(filename,conf):
    with app.app_context():
        success_count = 0
        error_count = 0
        sheet = pd.read_excel(filename)
        sheet = sheet.fillna('')
        Purch = []
        company = conf["company"]
        date = conf["date"]
        for row in sheet.itertuples():
            purchases= {}
            purchases['name']=row[conf["name"]+1]
            purchases['billvalue']=round(row[conf["billvalue"]+1])
            Purch.append(purchases)
        for x in Purch:
                print(x)
                custid = -1
                getcustid = MasterList.query.filter_by(name=x['name']).first()
                if getcustid is not None:
                    custid = getcustid.cust_id
                else:
                    error_count += 1
                    newError = ErrorPurchases(
                        name = x['name'],
                        interface_code = company.upper(),
                        billvalue = x['billvalue'],
                        date = date
                    ) 
                    db.session.add(newError)
                    db.session.commit()
                    continue
                purchases = PurchaseTable(
                    custid = custid,
                    interface_code = company.upper(),
                    billvalue = x['billvalue'],
                    date = date
                )
                db.session.add(purchases)
                success_count += 1
                db.session.commit()
        return (success_count,error_count)


def store(filename,company):
    with app.app_context():
        success_count = 0
        error_count = 0
        wb = op.load_workbook(filename)
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        Purchases = []
        for row in range(int(configs[company]['start']), sheet.max_row-int(configs[company]['end'])):
            purchase = {}
            for key,values in configs[company]['get'].items():
                if(values!=''):
                    # print(values+str(row))
                    purchase[key] = sheet[values+str(row)].value
                else:
                    purchase[key] = 0
            Purchases.append(purchase)

        for i in range(len(Purchases)):
            custid = -1
            getcustid = MasterList.query.filter_by(name=Purchases[i]['party']).first()
            if company=='marg':
                Purchases[i]['date'] = datetime.datetime.strptime(Purchases[i]['date'],'%d-%m-%y').strftime('%Y-%m-%d')
            if company=='cad':
                Purchases[i]['date'] = datetime.datetime.strptime(Purchases[i]['date'],'%d/%m/%Y').strftime('%Y-%m-%d')
            if getcustid is not None:
                custid = getcustid.cust_id   
            else:
                error_count += 1
                newError = ErrorPurchases(
                    name = Purchases[i]['party'],
                    interface_code = company.upper(),
                    billvalue = Purchases[i]['billvalue'],
                    billno = Purchases[i]['billno'],
                    date = Purchases[i]['date']
                ) 
                db.session.add(newError)
                db.session.commit()
                continue
            Purchases[i]['custid'] = custid
            Purchases[i].pop('party')
            Purchases[i].pop('tinno')
            Purchases[i]['interface_code']=company
            currentPurchase = Purchases[i]
            purchase = PurchaseTable(
                custid = currentPurchase['custid'],
                interface_code = currentPurchase['interface_code'].upper(),
                billvalue = currentPurchase['billvalue'],
                billno = currentPurchase['billno'],
                date = currentPurchase['date']
                )
            db.session.add(purchase)
            success_count += 1
            db.session.commit()
        return (success_count,error_count)


def main():
    app.run(debug=True)

if __name__ == '__main__':
    db.create_all()
    main()